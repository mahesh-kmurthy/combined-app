from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import yfinance as yf
import pandas as pd
import numpy as np
import io
import openpyxl
import os
import shutil
import traceback
import requests
import datetime

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------------------------------------------
# Portfolio Tracker Endpoints
# -------------------------------------------------------------

@app.get("/api/quotes")
def api_quotes(tickers: str = ""):
    if not tickers:
        raise HTTPException(status_code=400, detail="Missing tickers parameter")
        
    try:
        result = []
        finnhub_key = os.environ.get('FINNHUB_API_KEY')
        
        for symbol in tickers.split(','):
            price, change, pct_change, market_cap = 0, 0, 0, 0
            industry = 'N/A'
            name = ''
            finnhub_success = False
            
            if finnhub_key:
                try:
                    quote_url = f"https://finnhub.io/api/v1/quote?symbol={symbol}&token={finnhub_key}"
                    q_res = requests.get(quote_url, timeout=5)
                    
                    profile_url = f"https://finnhub.io/api/v1/stock/profile2?symbol={symbol}&token={finnhub_key}"
                    p_res = requests.get(profile_url, timeout=5)
                    
                    if q_res.status_code == 200 and p_res.status_code == 200:
                        q_data = q_res.json()
                        p_data = p_res.json()
                        
                        if q_data and q_data.get('c', 0) > 0:
                            price = q_data.get('c', 0)
                            change = q_data.get('d', 0)
                            if change is None: change = price - q_data.get('pc', 0)
                            pct_change = q_data.get('dp', 0)
                            if pct_change is None: pct_change = (change / q_data.get('pc', 1)) * 100 if q_data.get('pc') else 0
                            
                            if p_data:
                                industry = p_data.get('finnhubIndustry', 'N/A')
                                name = p_data.get('name', '')
                                raw_mc = p_data.get('marketCapitalization', 0)
                                market_cap = raw_mc * 1000000 if raw_mc else 0
                                
                            finnhub_success = True
                except Exception as e:
                    print(f"Finnhub Quote error for {symbol}: {e}")
            
            if finnhub_success:
                print(f"[Finnhub] Successfully fetched quote for {symbol}")
            else:
                print(f"[Yahoo Fallback] Fetching quote for {symbol} via yfinance")
                tkr = yf.Ticker(symbol)
                fi = tkr.fast_info
                try:
                    info = tkr.info
                    industry = info.get('industry', 'N/A')
                    name = info.get('longName', info.get('shortName', ''))
                except Exception:
                    industry = 'N/A'
                    name = ''
                    info = {}
                    
                # Extract basic fast_info fields
                try:
                    price = fi.last_price
                    prev = fi.previous_close
                    change = price - prev
                    pct_change = (change / prev * 100) if prev and prev > 0 else 0
                except Exception:
                    # Fallback if fast_info fails
                    price = 0
                    change = 0
                    pct_change = 0
                    
                market_cap = info.get('marketCap', 0)
                if not market_cap:
                    try:
                        market_cap = fi.market_cap
                    except Exception:
                        market_cap = 0
                
            result.append({
                "symbol": symbol,
                "name": name,
                "regularMarketPrice": float(price),
                "regularMarketChange": float(change),
                "regularMarketChangePercent": float(pct_change),
                "industry": industry,
                "marketCap": float(market_cap) if market_cap else 0
            })
            
        return {
            "quoteResponse": {
                "result": result
            }
        }
    except Exception as e:
        print("Error fetching quotes:")
        traceback.print_exc()
        return JSONResponse(status_code=500, content={'error': str(e)})

@app.get("/api/history")
def api_history(ticker: str = "VOO"):
    finnhub_key = os.environ.get('FINNHUB_API_KEY')
    
    if finnhub_key:
        try:
            end_dt = datetime.datetime.now()
            start_dt = end_dt - datetime.timedelta(days=3650) # 10 years
            
            fh_url = f"https://finnhub.io/api/v1/stock/candle?symbol={ticker}&resolution=D&from={int(start_dt.timestamp())}&to={int(end_dt.timestamp())}&token={finnhub_key}"
            fh_res = requests.get(fh_url, timeout=5)
            
            if fh_res.status_code == 200:
                fh_data = fh_res.json()
                if fh_data.get('s') == 'ok' and 't' in fh_data and 'c' in fh_data:
                    print(f"[Finnhub] Successfully fetched history for {ticker}")
                    return {
                        "chart": {
                            "result": [
                                {
                                    "timestamp": fh_data['t'],
                                    "indicators": {
                                        "quote": [
                                            { "close": fh_data['c'] }
                                        ]
                                    }
                                }
                            ]
                        }
                    }
        except Exception as e:
            print(f"Finnhub History error for {ticker}: {e}")
            
    # Fallback to yfinance
    print(f"[Yahoo Fallback] Fetching history for {ticker} via yfinance")
    try:
        tkr = yf.Ticker(ticker)
        # Fetch 10 years, daily
        hist = tkr.history(period="10y", interval="1d")
        
        if hist.empty:
            return JSONResponse(status_code=404, content={'error': 'No data found'})
            
        timestamps = []
        closes = []
        
        for index, row in hist.iterrows():
            timestamps.append(int(index.timestamp()))
            closes.append(float(row['Close']))
            
        return {
            "chart": {
                "result": [
                    {
                        "timestamp": timestamps,
                        "indicators": {
                            "quote": [
                                { "close": closes }
                            ]
                        }
                    }
                ]
            }
        }
    except Exception as e:
        print("Error fetching history:")
        traceback.print_exc()
        return JSONResponse(status_code=500, content={'error': str(e)})

# -------------------------------------------------------------
# Fundamental Analyzer Endpoints
# -------------------------------------------------------------

def safe_get(df, row_name, col_name, default=0):
    if row_name in df.index and col_name in df.columns:
        val = df.loc[row_name, col_name]
        if pd.isna(val) or np.isinf(val):
            return default
        return float(val)
    return default

def safe_div(num, den, default=0.0):
    if den == 0 or pd.isna(den) or np.isinf(den):
        return default
    val = num / den
    if pd.isna(val) or np.isinf(val):
        return default
    return float(val)

@app.get("/api/stock/{ticker}")
def get_stock_data(ticker: str):
    stock = yf.Ticker(ticker)
    
    # Financials (reverse rows to display logical top-to-bottom order)
    bs = stock.balance_sheet.iloc[::-1] if not stock.balance_sheet.empty else stock.balance_sheet
    is_ = stock.financials.iloc[::-1] if not stock.financials.empty else stock.financials
    cf = stock.cashflow.iloc[::-1] if not stock.cashflow.empty else stock.cashflow
    info = stock.info
    
    # Need 5 years of data. yfinance returns 4-5 years usually.
    # Convert dates to strings for JSON serializability
    cols = list(bs.columns)[:5]
    if not cols:
        raise HTTPException(status_code=404, detail="No financial data found")

    dates = [str(c).split(" ")[0] for c in cols]
    
    # We will compute metrics for the last 4 years using 5 years of data for YoY calculations
    calc_years = min(4, len(cols) - 1)
    if calc_years < 1:
        # fallback if not enough history
        calc_years = len(cols)
        
    metrics = []
    historical_pes = []
    
    # Fetch 10y historical prices for Average P/E calculation
    hist = stock.history(period="10y")
    if not hist.empty:
        hist.index = pd.to_datetime(hist.index).tz_localize(None)
    
    # Helpers
    current_price = info.get("currentPrice", info.get("regularMarketPrice", 0))
    shares_out = info.get("sharesOutstanding", 1) # Default to 1 to avoid div0
    
    curr_eps = 0
    curr_fcf_ps = 0
    curr_bvps = 0
    curr_pat_growth = 0
    
    # Iterate through the first `calc_years` (most recent to older)
    for i in range(calc_years):
        c_yr = cols[i]
        p_yr = cols[i+1] if i+1 < len(cols) else c_yr # Fallback to 0 growth if no previous year
        
        # Base Data from Income Statement
        net_income = safe_get(is_, "Net Income", c_yr)
        prev_net_income = safe_get(is_, "Net Income", p_yr)
        revenue = safe_get(is_, "Total Revenue", c_yr)
        prev_revenue = safe_get(is_, "Total Revenue", p_yr)
        ebit = safe_get(is_, "EBIT", c_yr)
        ebitda = safe_get(is_, "EBITDA", c_yr)
        prev_ebitda = safe_get(is_, "EBITDA", p_yr)
        tax_rate = safe_get(is_, "Tax Rate For Calcs", c_yr, 0.21)

        # Base Data from Balance Sheet
        total_assets = safe_get(bs, "Total Assets", c_yr)
        current_liabilities = safe_get(bs, "Current Liabilities", c_yr)
        total_debt = safe_get(bs, "Total Debt", c_yr)
        total_equity = safe_get(bs, "Stockholders Equity", c_yr, default=safe_get(bs, "Total Equity Gross Minority Interest", c_yr))
        cash_eq = safe_get(bs, "Cash And Cash Equivalents", c_yr)

        diluted_shares = safe_get(is_, "Diluted Average Shares", c_yr, shares_out)
        prev_diluted_shares = safe_get(is_, "Diluted Average Shares", p_yr, shares_out)

        # Base Data from Cash Flow
        ocf = safe_get(cf, "Operating Cash Flow", c_yr)
        prev_ocf = safe_get(cf, "Operating Cash Flow", p_yr)
        capex = safe_get(cf, "Capital Expenditure", c_yr)
        if capex < 0: capex = -capex
        fcf = ocf - capex
        
        prev_capex = safe_get(cf, "Capital Expenditure", p_yr)
        if prev_capex < 0: prev_capex = -prev_capex
        prev_fcf = prev_ocf - prev_capex

        # ---------------- CALCULATIONS ----------------
        # Summary
        market_cap = current_price * shares_out if i == 0 else None
        enterprise_value = total_equity + total_debt - cash_eq
        eps_diluted = safe_get(is_, "Diluted EPS", c_yr, safe_div(net_income, diluted_shares))
        bvps_diluted = safe_div(total_equity, diluted_shares)
        ocf_ps = safe_div(ocf, diluted_shares)
        fcf_ps = safe_div(fcf, diluted_shares)

        # Profitability
        op_margin = safe_div(safe_get(is_, "Operating Income", c_yr, ebit), revenue)
        ebitda_margin = safe_div(ebitda, revenue)
        net_profit_margin = safe_div(net_income, revenue)
        cfo_to_net_profit = safe_div(ocf, net_income)

        # Performance
        roe = safe_div(net_income, total_equity)
        capital_employed = total_assets - current_liabilities
        roce = safe_div(ebit, capital_employed)
        invested_capital = total_debt + total_equity
        nopat = ebit * (1 - tax_rate)
        roic = safe_div(nopat, invested_capital)

        # Growth Ratios (YoY)
        rev_growth = safe_div(revenue - prev_revenue, abs(prev_revenue)) if i+1 < len(cols) else 0
        ebitda_growth = safe_div(ebitda - prev_ebitda, abs(prev_ebitda)) if i+1 < len(cols) else 0
        pat_growth = safe_div(net_income - prev_net_income, abs(prev_net_income)) if i+1 < len(cols) else 0
        ocf_growth = safe_div(ocf - prev_ocf, abs(prev_ocf)) if i+1 < len(cols) else 0
        fcf_growth = safe_div(fcf - prev_fcf, abs(prev_fcf)) if i+1 < len(cols) else 0

        # Financial Stability
        debt_to_equity = safe_div(total_debt, total_equity)
        yoy_dilution = safe_div(diluted_shares - prev_diluted_shares, prev_diluted_shares) if i+1 < len(cols) else 0

        if i == 0:
            curr_eps = eps_diluted
            curr_fcf_ps = fcf_ps
            curr_bvps = bvps_diluted
            curr_pat_growth = pat_growth

        # Collect historical P/E for 5Y average
        if not hist.empty and eps_diluted and eps_diluted != 0:
            c_yr_naive = pd.to_datetime(c_yr).tz_localize(None)
            past_prices = hist.loc[:c_yr_naive]
            if not past_prices.empty:
                hist_price = float(past_prices["Close"].iloc[-1])
                historical_pes.append(hist_price / eps_diluted)

        metrics.append({
            "year": dates[i],
            "Summary": {
                "Market Cap": market_cap,
                "Enterprise Value": enterprise_value
            },
            "Profitability Ratios": {
                "Operating Margin": op_margin,
                "EBITDA Margin": ebitda_margin,
                "Net Profit Margin": net_profit_margin,
                "CFO / Net Profit": cfo_to_net_profit
            },
            "Performance Ratios": {
                "Return on Equity": roe,
                "Return on Capital Employed": roce,
                "Return on Invested Capital": roic
            },
            "Growth Ratios": {
                "Revenue Growth": rev_growth,
                "EBITDA Growth": ebitda_growth,
                "PAT Growth": pat_growth,
                "OCF Growth": ocf_growth,
                "FCF Growth": fcf_growth
            },
            "Financial Stability Ratios": {
                "Debt / Equity": debt_to_equity,
                "YoY Dilution": yoy_dilution
            }
        })

    # Valuation Ratios (Current Year Only)
    pe_ratio, peg_ratio, pb_ratio, ev_ebitda, p_fcf = 0, 0, 0, 0, 0
    if len(metrics) > 0 and current_price > 0:
        ev_current = metrics[0]["Summary"]["Enterprise Value"]
        ebitda_current = safe_get(is_, "EBITDA", cols[0])

        pe_ratio = safe_div(current_price, curr_eps)
        peg_ratio = safe_div(pe_ratio, curr_pat_growth * 100) if curr_pat_growth > 0.0 else 0
        pb_ratio = safe_div(current_price, curr_bvps)
        ev_ebitda = safe_div(ev_current, ebitda_current)
        p_fcf = safe_div(current_price, curr_fcf_ps)

    avg_pe_5y = 0.0
    valid_pes = [pe for pe in historical_pes if pe is not None and pe > 0]
    if len(valid_pes) > 0:
        avg_pe_5y = sum(valid_pes) / len(valid_pes)

    valuation_ratios = {
        "Price to Earnings": pe_ratio,
        "Average 5Y P/E": avg_pe_5y,
        "PEG": peg_ratio,
        "Price to Book": pb_ratio,
        "EV/EBITDA": ev_ebitda,
        "Price / Free Cash Flow": p_fcf
    }

    def format_df(df):
        df_json = {}
        avail_cols = [c for c in cols if c in df.columns]
        for row in df.index:
            row_data = {}
            for c in avail_cols:
                date_str = str(c).split(" ")[0]
                val = df.loc[row, c]
                if pd.isna(val) or np.isinf(val):
                    val = None
                row_data[date_str] = val
            df_json[str(row)] = row_data
        return df_json

    sales_growth_total = 0.0
    sg_count = 0
    for i in range(min(3, len(metrics))):
        st = metrics[i]["Growth Ratios"]["Revenue Growth"]
        if st is not None:
            sales_growth_total += float(st)
            sg_count += 1
            
    avg_3y_sales_growth = (sales_growth_total / sg_count) if sg_count > 0 else 0.0

    ttm_eps = info.get("trailingEps")
    if pd.isna(ttm_eps) or ttm_eps is None:
        ttm_eps = curr_eps
        
    valuation_model_params = {
        "ttm_eps": ttm_eps,
        "avg_3y_sales_growth": avg_3y_sales_growth
    }

    return {
        "ticker": ticker.upper(),
        "name": info.get("longName", info.get("shortName", "")),
        "current_price": current_price,
        "dates": dates,
        "metrics": metrics,
        "valuation": valuation_ratios,
        "model_params": valuation_model_params,
        "financials": {
            "balance_sheet": format_df(bs),
            "income_statement": format_df(is_),
            "cash_flow": format_df(cf)
        }
    }

# UPDATE FOR VERCEL SERVERLESS: Use /tmp for writable operations
TEMPLATE_PATH = "/tmp/user_template.xlsx"

@app.post("/api/template/upload")
async def upload_template(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    with open(TEMPLATE_PATH, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"message": "Template uploaded successfully"}

@app.get("/api/stock/{ticker}/export")
def export_stock_data(ticker: str):
    stock = yf.Ticker(ticker)
    bs = stock.balance_sheet
    is_ = stock.financials
    cf = stock.cashflow
    
    if bs.empty and is_.empty and cf.empty:
        raise HTTPException(status_code=404, detail="No financial data found")

    if os.path.exists(TEMPLATE_PATH):
        try:
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
        except Exception:
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()

    if not is_.empty:
        is_ = is_.iloc[::-1]
    if not bs.empty:
        bs = bs.iloc[::-1]
    if not cf.empty:
        cf = cf.iloc[::-1]

    def get_or_create_sheet(wb, sheet_name):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
            return ws
        else:
            if len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
                ws = wb.active
                ws.title = sheet_name
                return ws
            else:
                return wb.create_sheet(title=sheet_name)
    
    def write_df_to_ws(ws, df, title):
        start_row = 1
        
        ws.cell(row=start_row, column=1, value=title).font = openpyxl.styles.Font(bold=True)
        start_row += 1
        
        col_idx = 2
        for col in df.columns:
            date_str = str(col).split(" ")[0]
            ws.cell(row=start_row, column=col_idx, value=date_str).font = openpyxl.styles.Font(bold=True)
            col_idx += 1
        start_row += 1
        
        for index, row in df.iterrows():
            idx_str = str(index)
            ws.cell(row=start_row, column=1, value=idx_str)
            
            # Determine if this row should be scaled to millions
            scale = True
            lower_idx = idx_str.lower()
            if 'eps' in lower_idx or 'rate' in lower_idx or 'ratio' in lower_idx:
                scale = False

            col_idx = 2
            for col in df.columns:
                val = row[col]
                if pd.isna(val) or np.isinf(val):
                    val = ""
                else:
                    if scale and isinstance(val, (int, float)):
                        val = val / 1000000.0
                ws.cell(row=start_row, column=col_idx, value=val)
                col_idx += 1
            start_row += 1
            
        ws.column_dimensions['A'].width = 35

    if not is_.empty:
        ws_is = get_or_create_sheet(wb, "Income Statement")
        write_df_to_ws(ws_is, is_, "Income Statement (in Millions)")
    if not bs.empty:
        ws_bs = get_or_create_sheet(wb, "Balance Sheet")
        write_df_to_ws(ws_bs, bs, "Balance Sheet (in Millions)")
    if not cf.empty:
        ws_cf = get_or_create_sheet(wb, "Cash Flow")
        write_df_to_ws(ws_cf, cf, "Cash Flow (in Millions)")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="{ticker.upper()}_financials.xlsx"'
    }
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

# -------------------------------------------------------------
# Search Endpoint
# -------------------------------------------------------------

@app.get("/api/search")
def api_search(q: str):
    if not q:
        return {"result": []}
    try:
        url = f"https://query2.finance.yahoo.com/v1/finance/search?q={q}&quotesCount=5&newsCount=0"
        headers = {'User-Agent': 'Mozilla/5.0'}
        res = requests.get(url, headers=headers, timeout=5)
        if res.status_code == 200:
            data = res.json()
            quotes = data.get('quotes', [])
            return {"result": [{"symbol": item.get('symbol', ''), "name": item.get('shortname', item.get('longname', ''))} for item in quotes if item.get('quoteType') in ['EQUITY', 'ETF', 'MUTUALFUND']]}
        return {"result": []}
    except Exception as e:
        print(f"Search error: {e}")
        return {"result": []}

# Add static file mounting back for LOCAL testing
from fastapi.staticfiles import StaticFiles
app.mount("/", StaticFiles(directory=".", html=True), name="static")
